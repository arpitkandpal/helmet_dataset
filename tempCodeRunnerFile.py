import cv2
import torch
from ultralytics import YOLO
from datetime import datetime
import smtplib
from email.message import EmailMessage
import ssl
import os
import geocoder
from openpyxl import Workbook, load_workbook
import hashlib

# === CONFIGURATION ===
SENDER_EMAIL = "arpitkandpal10b@gmail.com"
APP_PASSWORD = "gmttegpte"
RECEIVER_EMAIL = "arpit.kandpal2025@gmail.com"
EXCEL_FILE = "no_helmet_log.xlsx"
ALERTS_FOLDER = "alerts"
os.makedirs(ALERTS_FOLDER, exist_ok=True)

# === LOAD MODEL ===
model = YOLO("runs/detect/train/weights/best.pt")
model.fuse()

# === TRACK SENT IMAGES ===
sent_hashes = set()

# === Excel setup ===
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.append(["Date", "Time", "Location", "Image"])
    wb.save(EXCEL_FILE)

# === Send Email Function ===
def send_email(img_path, location, time_str, date_str):
    try:
        msg = EmailMessage()
        msg["Subject"] = "⚠️ No Helmet Detected"
        msg["From"] = SENDER_EMAIL
        msg["To"] = RECEIVER_EMAIL
        msg.set_content(f"No helmet detected on {date_str} at {time_str}\nLocation: {location}")

        with open(img_path, "rb") as f:
            file_data = f.read()
            msg.add_attachment(file_data, maintype="image", subtype="jpeg", filename=os.path.basename(img_path))

        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ssl.create_default_context()) as smtp:
            smtp.login(SENDER_EMAIL, APP_PASSWORD)
            smtp.send_message(msg)
        print(f"[EMAIL SENT] {img_path}")
    except Exception as e:
        print(f"[EMAIL ERROR] {e}")

# === Location Function ===
def get_location():
    try:
        g = geocoder.ip("me")
        return g.city + ", " + g.country
    except:
        return "Unknown"

# === Hash Image Content ===
def hash_image(img):
    return hashlib.md5(img.tobytes()).hexdigest()

# === Start Webcam Detection ===
cap = cv2.VideoCapture(0)

while True:
    ret, frame = cap.read()
    if not ret:
        break

    results = model.predict(source=frame, save=False, imgsz=640, conf=0.5, device="cpu")[0]

    for box in results.boxes:
        cls_id = int(box.cls[0].item())
        label = results.names[cls_id]

        if label == "nohelmet":
            # Compute image hash to avoid resending
            image_hash = hash_image(frame)
            if image_hash in sent_hashes:
                continue
            sent_hashes.add(image_hash)

            now = datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H:%M:%S")
            location = get_location()

            filename = f"{ALERTS_FOLDER}/nohelmet_{now.strftime('%Y%m%d_%H%M%S')}.jpg"
            cv2.imwrite(filename, frame)

            # Send email
            send_email(filename, location, time_str, date_str)

            # Log to Excel
            if os.path.exists(EXCEL_FILE):
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
                ws.append([date_str, time_str, location, filename])
                wb.save(EXCEL_FILE)

        # Draw box
        cords = box.xyxy[0].tolist()
        cv2.rectangle(frame, (int(cords[0]), int(cords[1])), (int(cords[2]), int(cords[3])), (0, 0, 255), 2)
        cv2.putText(frame, label, (int(cords[0]), int(cords[1]) - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.8, (0, 0, 255), 2)

    cv2.imshow("Helmet Detection", frame)
    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()
